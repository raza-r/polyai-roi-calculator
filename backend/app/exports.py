import io
import csv
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML, CSS
from .models import DealInputs, Results


class ExcelExporter:
    def create_workbook(self, inputs: DealInputs, results: Results) -> bytes:
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_inputs_sheet(wb, inputs)
        self._create_intents_sheet(wb, inputs)
        self._create_calcs_sheet(wb, inputs, results)
        self._create_sensitivity_sheet(wb, results)
        self._create_scenarios_sheet(wb, results)
        self._create_notes_sheet(wb)
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    
    def _create_inputs_sheet(self, wb: Workbook, inputs: DealInputs):
        ws = wb.create_sheet("Inputs")
        
        # Headers
        ws['A1'] = "Parameter"
        ws['B1'] = "Value"
        ws['C1'] = "Description"
        
        # Data
        data = [
            ("Annual Calls", inputs.annual_calls, "Total annual call volume"),
            ("Agent Cost/Min (£)", inputs.agent_cost_per_min, "Agent cost per minute"),
            ("Telco Cost/Min (£)", inputs.telco_cost_per_min, "Telephony cost per minute"),
            ("PolyAI Cost/Min (£)", inputs.polyai_cost_per_min, "PolyAI cost per minute"),
            ("ACW Minutes", inputs.acw_minutes, "After-call work minutes"),
            ("Baseline Abandon Rate", inputs.baseline_abandon_rate, "Current abandon rate"),
            ("AI Abandon Rate", inputs.ai_abandon_rate, "Post-AI abandon rate"),
            ("Business Hours Only", inputs.business_hours_only, "Business hours vs 24/7"),
            ("Night Fraction", inputs.night_fraction, "Fraction of calls outside business hours"),
            ("Inflation Rate", inputs.inflation, "Annual inflation rate"),
            ("Volume Growth", inputs.volume_growth, "Annual volume growth rate"),
            ("Discount Rate", inputs.discount_rate, "Discount rate for NPV"),
            ("Risk Adjustment", inputs.risk_adjustment, "Containment risk adjustment")
        ]
        
        for i, (param, value, desc) in enumerate(data, 2):
            ws[f'A{i}'] = param
            ws[f'B{i}'] = value
            ws[f'C{i}'] = desc
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_intents_sheet(self, wb: Workbook, inputs: DealInputs):
        ws = wb.create_sheet("Intents")
        
        # Headers
        headers = ["Intent Name", "Volume Share", "Avg Minutes", "Containment M0", 
                  "Containment M3", "Handoff Minutes", "Revenue/Abandon"]
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header)
        
        # Data
        for i, intent in enumerate(inputs.intents, 2):
            ws.cell(row=i, column=1, value=intent.name)
            ws.cell(row=i, column=2, value=intent.volume_share)
            ws.cell(row=i, column=3, value=intent.avg_minutes)
            ws.cell(row=i, column=4, value=intent.containment_m0)
            ws.cell(row=i, column=5, value=intent.containment_m3)
            ws.cell(row=i, column=6, value=intent.handoff_minutes)
            ws.cell(row=i, column=7, value=intent.revenue_per_abandon or "")
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_calcs_sheet(self, wb: Workbook, inputs: DealInputs, results: Results):
        ws = wb.create_sheet("Calcs_5Y")
        
        # Headers
        headers = ["Year", "Baseline Minutes", "Automated Minutes", "Handoff Minutes", 
                  "Human Minutes", "Baseline Cost (£)", "AI Cost (£)", "Ops Savings (£)",
                  "Revenue Retained (£)", "Total Value (£)", "Cumulative Value (£)", 
                  "Discounted Value (£)"]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header)
        
        # Data
        for i, year_result in enumerate(results.yearly, 2):
            ws.cell(row=i, column=1, value=year_result.year)
            ws.cell(row=i, column=2, value=year_result.baseline_minutes)
            ws.cell(row=i, column=3, value=year_result.automated_minutes)
            ws.cell(row=i, column=4, value=year_result.handoff_minutes)
            ws.cell(row=i, column=5, value=year_result.human_minutes)
            ws.cell(row=i, column=6, value=year_result.baseline_cost)
            ws.cell(row=i, column=7, value=year_result.ai_cost)
            ws.cell(row=i, column=8, value=year_result.ops_savings)
            ws.cell(row=i, column=9, value=year_result.revenue_retained)
            ws.cell(row=i, column=10, value=year_result.total_value)
            ws.cell(row=i, column=11, value=year_result.cumulative_value)
            ws.cell(row=i, column=12, value=year_result.discounted_value)
        
        # Summary metrics
        summary_row = len(results.yearly) + 3
        ws.cell(row=summary_row, column=1, value="Summary Metrics")
        ws.cell(row=summary_row + 1, column=1, value="5Y ROI (%)")
        ws.cell(row=summary_row + 1, column=2, value=results.roi_5y)
        ws.cell(row=summary_row + 2, column=1, value="5Y NPV (£)")
        ws.cell(row=summary_row + 2, column=2, value=results.npv_5y)
        ws.cell(row=summary_row + 3, column=1, value="Payback (months)")
        ws.cell(row=summary_row + 3, column=2, value=results.payback_months or "No payback")
        
        # Style
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 15)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_sensitivity_sheet(self, wb: Workbook, results: Results):
        ws = wb.create_sheet("Sensitivity")
        
        ws['A1'] = "Driver"
        ws['B1'] = "Impact on NPV (£)"
        
        for i, (driver, impact) in enumerate(results.tornado, 2):
            ws[f'A{i}'] = driver
            ws[f'B{i}'] = impact
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['B1'].font = header_font
        ws['B1'].fill = header_fill
    
    def _create_scenarios_sheet(self, wb: Workbook, results: Results):
        ws = wb.create_sheet("Scenarios")
        
        ws['A1'] = "Scenario"
        ws['B1'] = "NPV (£)"
        
        ws['A2'] = "P10 (Pessimistic)"
        ws['B2'] = results.p10_p50_p90["p10"]
        
        ws['A3'] = "P50 (Base Case)"
        ws['B3'] = results.p10_p50_p90["p50"]
        
        ws['A4'] = "P90 (Optimistic)"
        ws['B4'] = results.p10_p50_p90["p90"]
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['B1'].font = header_font
        ws['B1'].fill = header_fill
    
    def _create_notes_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Notes")
        
        notes = [
            "ROI Calculator - Assumptions and Disclaimers",
            "",
            "This analysis is illustrative and based on assumptions that may vary by deployment.",
            "",
            "Key Assumptions:",
            "- Containment rates improve linearly from M0 to M3 over 3 months",
            "- Costs include agent time, ACW, telephony, and PolyAI usage",
            "- Revenue impact calculated from abandon rate reduction",
            "- All values in GBP unless specified",
            "",
            "Disclaimers:",
            "- Not official PolyAI pricing",
            "- Results depend on actual deployment characteristics",
            "- Sensitivity analysis shows impact of ±20% parameter changes",
            "- P10/P50/P90 scenarios use triangular distribution approximation"
        ]
        
        for i, note in enumerate(notes, 1):
            ws[f'A{i}'] = note
        
        # Auto-size column
        ws.column_dimensions['A'].width = 80


class PDFExporter:
    def create_report(self, inputs: DealInputs, results: Results) -> bytes:
        html_content = self._generate_html(inputs, results)
        css_content = self._generate_css()
        
        html_doc = HTML(string=html_content)
        css_doc = CSS(string=css_content)
        
        pdf_buffer = io.BytesIO()
        html_doc.write_pdf(pdf_buffer, stylesheets=[css_doc])
        
        return pdf_buffer.getvalue()
    
    def _generate_html(self, inputs: DealInputs, results: Results) -> str:
        # Summary KPIs
        payback_str = f"{results.payback_months:.1f} months" if results.payback_months else "No payback"
        
        # Ops vs Revenue split
        ops_pct = results.ops_vs_revenue_split.get("ops_savings", 0)
        revenue_pct = results.ops_vs_revenue_split.get("revenue_retained", 0)
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>ROI Analysis Report</title>
        </head>
        <body>
            <!-- Page 1: Summary -->
            <div class="page">
                <h1>Voice AI ROI Analysis</h1>
                
                <div class="kpi-grid">
                    <div class="kpi-card">
                        <h3>5-Year Total Value</h3>
                        <div class="kpi-value">£{sum(yr.total_value for yr in results.yearly):,.0f}</div>
                    </div>
                    <div class="kpi-card">
                        <h3>Payback Period</h3>
                        <div class="kpi-value">{payback_str}</div>
                    </div>
                    <div class="kpi-card">
                        <h3>5-Year NPV</h3>
                        <div class="kpi-value">£{results.npv_5y:,.0f}</div>
                    </div>
                    <div class="kpi-card">
                        <h3>5-Year ROI</h3>
                        <div class="kpi-value">{results.roi_5y:.1f}%</div>
                    </div>
                </div>
                
                <h2>Value Driver Breakdown</h2>
                <div class="split-chart">
                    <div class="split-item">
                        <span>Operational Savings: {ops_pct:.1f}%</span>
                    </div>
                    <div class="split-item">
                        <span>Revenue Retained: {revenue_pct:.1f}%</span>
                    </div>
                </div>
                
                <h2>5-Year Financial Projection</h2>
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>Year</th>
                            <th>Baseline Cost (£)</th>
                            <th>AI Cost (£)</th>
                            <th>Ops Savings (£)</th>
                            <th>Revenue Retained (£)</th>
                            <th>Total Value (£)</th>
                        </tr>
                    </thead>
                    <tbody>
        """
        
        for yr in results.yearly:
            html += f"""
                        <tr>
                            <td>Year {yr.year}</td>
                            <td>£{yr.baseline_cost:,.0f}</td>
                            <td>£{yr.ai_cost:,.0f}</td>
                            <td>£{yr.ops_savings:,.0f}</td>
                            <td>£{yr.revenue_retained:,.0f}</td>
                            <td>£{yr.total_value:,.0f}</td>
                        </tr>
            """
        
        html += """
                    </tbody>
                </table>
            </div>
            
            <!-- Page 2: Sensitivity Analysis -->
            <div class="page">
                <h1>Sensitivity Analysis</h1>
                
                <h2>Tornado Chart - Top Value Drivers</h2>
                <div class="tornado-chart">
        """
        
        for driver, impact in results.tornado:
            html += f"""
                    <div class="tornado-bar">
                        <span class="driver-name">{driver}</span>
                        <div class="bar-container">
                            <div class="bar" style="width: {min(100, abs(impact)/10000):.1f}%"></div>
                        </div>
                        <span class="impact-value">±£{abs(impact):,.0f}</span>
                    </div>
            """
        
        html += f"""
                </div>
                
                <h2>Scenario Analysis</h2>
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>Scenario</th>
                            <th>5-Year NPV (£)</th>
                            <th>vs Base Case</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr>
                            <td>P10 (Pessimistic)</td>
                            <td>£{results.p10_p50_p90['p10']:,.0f}</td>
                            <td>{((results.p10_p50_p90['p10'] / results.p10_p50_p90['p50'] - 1) * 100):+.1f}%</td>
                        </tr>
                        <tr class="base-case">
                            <td>P50 (Base Case)</td>
                            <td>£{results.p10_p50_p90['p50']:,.0f}</td>
                            <td>Base</td>
                        </tr>
                        <tr>
                            <td>P90 (Optimistic)</td>
                            <td>£{results.p10_p50_p90['p90']:,.0f}</td>
                            <td>{((results.p10_p50_p90['p90'] / results.p10_p50_p90['p50'] - 1) * 100):+.1f}%</td>
                        </tr>
                    </tbody>
                </table>
            </div>
            
            <!-- Page 3: Assumptions -->
            <div class="page">
                <h1>Key Assumptions</h1>
                
                <h2>Business Parameters</h2>
                <table class="data-table">
                    <tbody>
                        <tr><td>Annual Call Volume</td><td>{inputs.annual_calls:,}</td></tr>
                        <tr><td>Agent Cost per Minute</td><td>£{inputs.agent_cost_per_min:.2f}</td></tr>
                        <tr><td>PolyAI Cost per Minute</td><td>£{inputs.polyai_cost_per_min:.2f}</td></tr>
                        <tr><td>Baseline Abandon Rate</td><td>{inputs.baseline_abandon_rate:.1%}</td></tr>
                        <tr><td>AI Abandon Rate</td><td>{inputs.ai_abandon_rate:.1%}</td></tr>
                        <tr><td>Volume Growth (Annual)</td><td>{inputs.volume_growth:.1%}</td></tr>
                        <tr><td>Discount Rate</td><td>{inputs.discount_rate:.1%}</td></tr>
                    </tbody>
                </table>
                
                <h2>Intent Mix</h2>
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>Intent</th>
                            <th>Volume Share</th>
                            <th>Avg Minutes</th>
                            <th>M3 Containment</th>
                        </tr>
                    </thead>
                    <tbody>
        """
        
        for intent in inputs.intents:
            html += f"""
                        <tr>
                            <td>{intent.name}</td>
                            <td>{intent.volume_share:.1%}</td>
                            <td>{intent.avg_minutes:.1f}</td>
                            <td>{intent.containment_m3:.1%}</td>
                        </tr>
            """
        
        html += """
                    </tbody>
                </table>
                
                <div class="disclaimer">
                    <h3>Disclaimers</h3>
                    <ul>
                        <li>This analysis is illustrative and not official PolyAI pricing</li>
                        <li>Results depend on actual deployment characteristics</li>
                        <li>Containment rates assume linear ramp from M0 to M3 over 3 months</li>
                        <li>Usage pricing is per-minute as deployed</li>
                        <li>All assumptions are editable in the calculator</li>
                    </ul>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def _generate_css(self) -> str:
        return """
        @page {
            size: A4;
            margin: 2cm;
        }
        
        body {
            font-family: Arial, sans-serif;
            font-size: 11pt;
            line-height: 1.4;
            color: #333;
        }
        
        .page {
            page-break-after: always;
        }
        
        .page:last-child {
            page-break-after: avoid;
        }
        
        h1 {
            color: #2c3e50;
            border-bottom: 2px solid #3498db;
            padding-bottom: 10px;
            margin-bottom: 20px;
        }
        
        h2 {
            color: #34495e;
            margin-top: 25px;
            margin-bottom: 15px;
        }
        
        .kpi-grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 15px;
            margin-bottom: 25px;
        }
        
        .kpi-card {
            border: 1px solid #ddd;
            padding: 15px;
            text-align: center;
            background-color: #f8f9fa;
        }
        
        .kpi-card h3 {
            margin: 0 0 10px 0;
            font-size: 12pt;
            color: #5a6c7d;
        }
        
        .kpi-value {
            font-size: 18pt;
            font-weight: bold;
            color: #2c3e50;
        }
        
        .data-table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
        }
        
        .data-table th,
        .data-table td {
            padding: 8px;
            text-align: left;
            border: 1px solid #ddd;
        }
        
        .data-table th {
            background-color: #f8f9fa;
            font-weight: bold;
        }
        
        .data-table .base-case {
            background-color: #e8f5e8;
        }
        
        .tornado-chart {
            margin: 20px 0;
        }
        
        .tornado-bar {
            display: flex;
            align-items: center;
            margin-bottom: 10px;
            font-size: 10pt;
        }
        
        .driver-name {
            width: 150px;
            flex-shrink: 0;
        }
        
        .bar-container {
            flex-grow: 1;
            height: 20px;
            background-color: #f0f0f0;
            margin: 0 10px;
            position: relative;
        }
        
        .bar {
            height: 100%;
            background-color: #3498db;
        }
        
        .impact-value {
            width: 80px;
            text-align: right;
            flex-shrink: 0;
        }
        
        .split-chart {
            margin: 15px 0;
        }
        
        .split-item {
            padding: 10px;
            margin: 5px 0;
            background-color: #f8f9fa;
            border-left: 4px solid #3498db;
        }
        
        .disclaimer {
            margin-top: 30px;
            padding: 15px;
            background-color: #f8f9fa;
            border: 1px solid #ddd;
        }
        
        .disclaimer h3 {
            margin-top: 0;
            color: #e74c3c;
        }
        
        .disclaimer ul {
            margin-bottom: 0;
        }
        """


class CSVExporter:
    def create_csv(self, inputs: DealInputs, results: Results) -> str:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            "Year", "Baseline_Minutes", "Automated_Minutes", "Handoff_Minutes",
            "Human_Minutes", "Baseline_Cost_GBP", "AI_Cost_GBP", "Ops_Savings_GBP",
            "Revenue_Retained_GBP", "Total_Value_GBP", "Cumulative_Value_GBP",
            "Discounted_Value_GBP", "Annual_Calls", "Agent_Cost_Per_Min",
            "PolyAI_Cost_Per_Min", "Payback_Months", "ROI_5Y_Percent", "NPV_5Y_GBP"
        ])
        
        # Data rows
        for yr in results.yearly:
            writer.writerow([
                yr.year,
                round(yr.baseline_minutes, 0),
                round(yr.automated_minutes, 0),
                round(yr.handoff_minutes, 0),
                round(yr.human_minutes, 0),
                round(yr.baseline_cost, 2),
                round(yr.ai_cost, 2),
                round(yr.ops_savings, 2),
                round(yr.revenue_retained, 2),
                round(yr.total_value, 2),
                round(yr.cumulative_value, 2),
                round(yr.discounted_value, 2),
                inputs.annual_calls * ((1 + inputs.volume_growth) ** yr.year),
                inputs.agent_cost_per_min,
                inputs.polyai_cost_per_min,
                results.payback_months,
                round(results.roi_5y, 2),
                round(results.npv_5y, 2)
            ])
        
        return output.getvalue()